#coding=utf8
import itchat
from itchat.content import *
import os
import uuid
import hashlib
import time, datetime
from itchat.content import *
import json,re
import requests
 
from openpyxl import  Workbook 
from openpyxl  import load_workbook
from openpyxl.styles import numbers


#文件存储位置
def mkdir(path):
    isExists = os.path.exists(path)
    if not isExists:
        os.makedirs(path)
        return True
    else:
        return False

def mkxls(date):
    filename = date+'.xlsx'
    path = 'D://tempfortest/hello/'+ filename
    
    isExists = os.path.exists(path)
    if not isExists:
        print('yes')
        # 实例化
        wb = Workbook()
        # 激活 worksheet
        ws = wb.active
        
        ws.append(['微信昵称','客户名称', '产品名称', '金额','期数','提交时间'])
        col_d = ws.column_dimensions['D']
        col_d.number_format = numbers.FORMAT_NUMBER_00
        col_e = ws.column_dimensions['E']
        col_e.number_format = numbers.FORMAT_NUMBER_00
        col_f = ws.column_dimensions['F']
        col_f.number_format = numbers.FORMAT_DATE_DATETIME
        wb.save(path)
        print('yes')
        return True
    else:
        return False


 
def is_number(s):
    try:
        float(s)
        return True
    except ValueError:
        pass

    try:
        import unicodedata
        unicodedata.numeric(s)
        return True
    except (TypeError, ValueError):
        pass
    return False
 


def verify_group_msg(contentlist):
    if len(contentlist) !=4 :
        print(len(contentlist))
        return False
    else:
        if not is_number(contentlist[2]) :
            return False
        else:
            if not is_number(contentlist[3]) :
                return False
            else:
                return True

def reply(content):
    print('in')
    itchat.send_msg(content, chatroom_ids[0])
        
#保存群消息
def save_group_msg(msg):
    msg_id = msg['MsgId']
    msg_from_user = msg['ActualNickName']
    msg_content = msg['Content']
    msg_create_time = msg['CreateTime']
    msg_type = msg['Type']
    #msg_group_nickname = msg['User']['NickName']
    # 使用time
    timeArray = time.localtime(msg_create_time)
    otherStyleTime = time.strftime("%Y-%m-%d %H:%M:%S", timeArray)

    contentlist = msg_content.split()

    if not verify_group_msg(contentlist):
        reply('登记失败')
        return
    else:
        date = time.strftime("%Y%m%d", timeArray)
        mkxls(date)
        filename = date+'.xlsx'
        path = 'D://tempfortest/hello/'+ filename
        # 设置文件 mingc
        wb = load_workbook(path)
        ws = wb.active
        ws.append([msg_from_user,contentlist[0],contentlist[1],contentlist[2],contentlist[3],otherStyleTime])
        #ws.append(['微信昵称','客户名称', '产品名称', '金额','几期','提交时间'])
        wb.save(path)
        reply('登记成功')
    return 


chatroom_ids = []

@itchat.msg_register(itchat.content.TEXT, isGroupChat=True)  #注册对文本消息进行监听，对群聊进行监听
def print_content(msg):
    
    msg_from_user = msg['ActualNickName']
    msg_content = msg['Content']
    msg_chatroom_id = msg['User']['UserName']
    print('msg_chatroom_id:',msg_chatroom_id)
    print("群聊信息: ",msg_from_user, msg_content)

    if not msg_chatroom_id in chatroom_ids:
        return
    
    save_group_msg(msg)

   

# 轮询任务
def start_schedule():
    sched.add_job(logout, 'interval', minutes=1)
    sched.start()

def lc():
    print('login')
    print (time.strftime('%H:%M:%S',time.localtime(time.time())))
def ec():
    print('exit')
    print (time.strftime('%H:%M:%S',time.localtime(time.time())))


if __name__=="__main__":

    itchat.auto_login(hotReload=True,enableCmdQR=False,loginCallback=lc, exitCallback=ec)
    chat_rooms = itchat.search_chatrooms(name='test')
    if len(chat_rooms) > 0:
        print('id:',chat_rooms[0]['UserName'])
        chatroom_ids.append(chat_rooms[0]['UserName'])
        print(chatroom_ids[0])
    itchat.run()


